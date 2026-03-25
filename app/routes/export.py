import csv
import io
import json
from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.models import ExtractionTask, Business

export_bp = Blueprint('export', __name__)

ALLOWED_EXPORT_FIELDS = [
    'name', 'email', 'phone', 'website', 'location', 'owner',
]
FIELD_LABELS = {
    'name': 'Name',
    'email': 'Email',
    'phone': 'Phone',
    'website': 'Website',
    'location': 'Location',
    'owner': 'Owner',
}
DEFAULT_EXPORT_FIELDS = [
    'name', 'email', 'phone', 'website', 'location', 'owner',
]


def _normalize_fields(raw_fields):
    if isinstance(raw_fields, str):
        candidates = [v.strip().lower() for v in raw_fields.split(',') if v and v.strip()]
    elif isinstance(raw_fields, (list, tuple, set)):
        candidates = [str(v).strip().lower() for v in raw_fields if str(v).strip()]
    else:
        candidates = []

    normalized = []
    for field in candidates:
        if field in ALLOWED_EXPORT_FIELDS and field not in normalized:
            normalized.append(field)
    return normalized


def _get_task_selected_fields(task):
    if not task.selected_fields:
        return DEFAULT_EXPORT_FIELDS.copy()
    try:
        parsed = json.loads(task.selected_fields)
    except (TypeError, json.JSONDecodeError):
        return DEFAULT_EXPORT_FIELDS.copy()

    normalized = _normalize_fields(parsed)
    return normalized or DEFAULT_EXPORT_FIELDS.copy()


def _resolve_export_fields(task):
    query_fields = request.args.get('fields')
    if query_fields:
        normalized = _normalize_fields(query_fields)
        if normalized:
            return normalized
    return _get_task_selected_fields(task)

@export_bp.route('/csv/<int:task_id>', methods=['GET'])
@login_required
def export_csv(task_id):
    task = ExtractionTask.query.get_or_404(task_id)
    businesses = task.businesses
    export_fields = _resolve_export_fields(task)

    si = io.StringIO()
    writer = csv.DictWriter(si, fieldnames=export_fields)
    writer.writeheader()
    for b in businesses:
        row = b.to_dict()
        writer.writerow({field: row.get(field) for field in export_fields})

    output = io.BytesIO()
    output.write(si.getvalue().encode('utf-8'))
    output.seek(0)

    return send_file(
        output,
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'task_{task_id}_results.csv'
    )


@export_bp.route('/excel/<int:task_id>', methods=['GET'])
@login_required
def export_excel(task_id):
    task = ExtractionTask.query.get_or_404(task_id)
    businesses = task.businesses
    export_fields = _resolve_export_fields(task)

    wb = Workbook()
    ws = wb.active
    ws.title = f'Task {task_id} Results'

    # Header styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    # Write headers
    headers = [FIELD_LABELS.get(field, field.title()) for field in export_fields]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Write data rows
    for row_idx, b in enumerate(businesses, 2):
        data = b.to_dict()
        for col_idx, field in enumerate(export_fields, 1):
            value = data.get(field)
            if isinstance(value, (dict, list)):
                value = json.dumps(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_length = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'task_{task_id}_results.xlsx'
    )
